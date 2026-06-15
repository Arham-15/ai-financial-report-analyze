import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from processor import process_financials
from ai_summary import generate_summary
import os

def export_to_excel():
    df = process_financials()
    summary = generate_summary()

    wb = openpyxl.Workbook()

    # ─── COLORS ───
    HEADER_COLOR = "1F3864"   # dark blue
    TESLA_COLOR  = "F4CCCC"   # light red
    MSFT_COLOR   = "CFE2F3"   # light blue
    APPLE_COLOR  = "D9D9D9"   # light gray
    TITLE_COLOR  = "2E75B6"   # medium blue

    header_font   = Font(bold=True, color="FFFFFF", size=11)
    title_font    = Font(bold=True, color="FFFFFF", size=14)
    company_fonts = {
        "Tesla":     Font(bold=True, color="CC0000"),
        "Microsoft": Font(bold=True, color="0078D4"),
        "Apple":     Font(bold=True, color="555555"),
    }
    center = Alignment(horizontal="center", vertical="center")
    thin   = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin")
    )

    def style_cell(cell, bold=False, color=None, bg=None, align=True):
        if bold:
            cell.font = Font(bold=True)
        if color:
            cell.font = Font(bold=bold, color=color)
        if bg:
            cell.fill = PatternFill("solid", fgColor=bg)
        if align:
            cell.alignment = center
        cell.border = thin

    # SHEET 1 — Financial Metrics
    ws1 = wb.active
    ws1.title = "Financial Metrics"

    # Title row
    ws1.merge_cells("A1:F1")
    ws1["A1"] = "FY2025 Financial Report — Tesla vs Microsoft vs Apple"
    ws1["A1"].font = title_font
    ws1["A1"].fill = PatternFill("solid", fgColor=TITLE_COLOR)
    ws1["A1"].alignment = center

    # Headers
    headers = ["Company", "Revenue (B)", "Net Income (B)",
               "Total Assets (B)", "Total Liabilities (B)", "Cash Flow Ops (B)"]
    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=2, column=col, value=h)
        cell.font = header_font
        cell.fill = PatternFill("solid", fgColor=HEADER_COLOR)
        cell.alignment = center
        cell.border = thin

    # Data rows
    row_colors = {"Tesla": TESLA_COLOR, "Microsoft": MSFT_COLOR, "Apple": APPLE_COLOR}
    for row_idx, company in enumerate(df.index, 3):
        data = [company] + df.loc[company].tolist()[:5]
        for col_idx, value in enumerate(data, 1):
            cell = ws1.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = PatternFill("solid", fgColor=row_colors[company])
            cell.alignment = center
            cell.border = thin
            if col_idx == 1:
                cell.font = company_fonts[company]

    # Column widths
    for col in range(1, 7):
        ws1.column_dimensions[get_column_letter(col)].width = 22
    ws1.row_dimensions[1].height = 30
    ws1.row_dimensions[2].height = 20

    # SHEET 2 — Ratios
    
    ws2 = wb.create_sheet("Ratios")

    ws2.merge_cells("A1:C1")
    ws2["A1"] = "FY2025 Financial Ratios"
    ws2["A1"].font = title_font
    ws2["A1"].fill = PatternFill("solid", fgColor=TITLE_COLOR)
    ws2["A1"].alignment = center

    ratio_headers = ["Company", "Profit Margin (%)", "Debt Ratio (%)"]
    for col, h in enumerate(ratio_headers, 1):
        cell = ws2.cell(row=2, column=col, value=h)
        cell.font = header_font
        cell.fill = PatternFill("solid", fgColor=HEADER_COLOR)
        cell.alignment = center
        cell.border = thin

    for row_idx, company in enumerate(df.index, 3):
        row_data = [company,
                    df.loc[company]["Profit Margin (%)"],
                    df.loc[company]["Debt Ratio (%)"]]
        for col_idx, value in enumerate(row_data, 1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = PatternFill("solid", fgColor=row_colors[company])
            cell.alignment = center
            cell.border = thin
            if col_idx == 1:
                cell.font = company_fonts[company]

    for col in range(1, 4):
        ws2.column_dimensions[get_column_letter(col)].width = 22
    ws2.row_dimensions[1].height = 30

    
    # SHEET 3 — AI Summary
    
    ws3 = wb.create_sheet("AI Summary")

    ws3.merge_cells("A1:D1")
    ws3["A1"] = "AI Generated Financial Summary"
    ws3["A1"].font = title_font
    ws3["A1"].fill = PatternFill("solid", fgColor=TITLE_COLOR)
    ws3["A1"].alignment = center
    ws3.row_dimensions[1].height = 30

    ws3.merge_cells("A2:D12")
    ws3["A2"] = summary
    ws3["A2"].alignment = Alignment(wrap_text=True, vertical="top")
    ws3["A2"].font = Font(size=11)
    ws3.row_dimensions[2].height = 200

    for col in range(1, 5):
        ws3.column_dimensions[get_column_letter(col)].width = 30

    
    # SHEET 4 — Charts
    
    ws4 = wb.create_sheet("Charts")

    ws4.merge_cells("A1:N1")
    ws4["A1"] = "FY2025 Financial Charts"
    ws4["A1"].font = title_font
    ws4["A1"].fill = PatternFill("solid", fgColor=TITLE_COLOR)
    ws4["A1"].alignment = center
    ws4.row_dimensions[1].height = 30

    chart_files = [
        ("charts/revenue.png",       "A2"),
        ("charts/net_income.png",    "H2"),
        ("charts/cash_flow.png",     "A22"),
        ("charts/margin_vs_debt.png","H22"),
    ]

    for chart_path, anchor in chart_files:
        if os.path.exists(chart_path):
            img = Image(chart_path)
            img.width  = 400
            img.height = 250
            ws4.add_image(img, anchor)

    # ── Save ──
    output_path = "financial_report.xlsx"
    wb.save(output_path)
    print(f"Excel report saved: {output_path}")

if __name__ == "__main__":
    export_to_excel()